import streamlit as st
import pandas as pd
from io import BytesIO
import openpyxl

st.set_page_config(page_title="Multi-Sheet Data Mapper", layout="wide")
st.title("Excel Multi-Sheet & Filter Tool (Formatting Safe)")

# Sidebar - File Upload
st.sidebar.header("📂 Upload Files")
main_file = st.sidebar.file_uploader("Upload Main.xlsx (Copy)", type=['xlsx'])
trans_file = st.sidebar.file_uploader("Upload Transport.xlsx (Paste)", type=['xlsx'])

# දත්ත ගබඩා කිරීමට තාවකාලික Variables
df_main = None
df_trans = None

col_a, col_b = st.columns(2)

# --- Main File එකේ Sheet එක සහ Heading Row එක තේරීම ---
with col_a:
    if main_file:
        st.subheader("1. Source Sheet")
        xl_main = pd.ExcelFile(main_file)
        sheet_main = st.selectbox("Main file එකේ වැඩ කළ යුතු Sheet එක තෝරන්න", xl_main.sheet_names, key="main_sheet")
        header_row_main = st.number_input("Main File හි Heading එක ඇති පේළිය (Row Index 0 සිට)", min_value=0, max_value=10, value=0, key="main_header")
        df_main = pd.read_excel(main_file, sheet_name=sheet_main, header=header_row_main)
        st.write(f"දත්ත පේළි {len(df_main)} ක් හමු වුණා.")

# --- Transport File එකේ Sheet එක සහ Heading Row එක තේරීම ---
with col_b:
    if trans_file:
        st.subheader("2. Destination Sheet")
        xl_trans = pd.ExcelFile(trans_file)
        sheet_trans = st.selectbox("Transport file එකේ වැඩ කළ යුතු Sheet එක තෝරන්න", xl_trans.sheet_names, key="trans_sheet")
        header_row_trans = st.number_input("Transport File හි Heading එක ඇති පේළිය (Row Index 0 සිට)", min_value=0, max_value=10, value=0, key="trans_header")
        df_trans = pd.read_excel(trans_file, sheet_name=sheet_trans, header=header_row_trans)
        st.write(f"Heading {len(df_trans.columns)} ක් හමු වුණා.")

st.divider()

if df_main is not None and df_trans is not None:
    # --- Filter Section ---
    st.subheader("🎯 3. Row Filter")
    enable_filter = st.checkbox("දත්ත Filter කිරීමට මෙය Tick කරන්න")
    active_df = df_main.copy()
    if enable_filter:
        f_col1, f_col2 = st.columns(2)
        with f_col1:
            filter_col = st.selectbox("Filter කළ යුතු Column එක", df_main.columns)
        with f_col2:
            unique_vals = df_main[filter_col].unique().tolist()
            selected_vals = st.multiselect(f"'{filter_col}' හි අගයන් තෝරන්න", unique_vals)
        if selected_vals:
            active_df = df_main[df_main[filter_col].isin(selected_vals)]
            st.info(f"Filter කළ පේළි ගණන: {len(active_df)}")

    st.divider()

    # --- Mapping Section ---
    st.subheader("🔗 4. Map Columns")
    mapping_dict = {}
    h1, h2, h3 = st.columns([0.5, 2, 2])
    h1.write("**Select**")
    h2.write(f"**Source ({sheet_main})**")
    h3.write(f"**Destination ({sheet_trans})**")

    for i, m_col in enumerate(active_df.columns):
        c1, c2, c3 = st.columns([0.5, 2, 2])
        with c1:
            checked = st.checkbox("", key=f"chk_{i}")
        with c2:
            st.info(m_col)
        with c3:
            if checked:
                target = st.selectbox(f"Map {i}", df_trans.columns, key=f"tar_{i}", label_visibility="collapsed")
                mapping_dict[m_col] = target

    # --- Execute (Safe Update Logic) ---
    if st.button("🚀 Process & Download", use_container_width=True):
        if mapping_dict:
            try:
                # පවතින Excel එක Open කිරීම (Formatting ආරක්ෂා කිරීමට)
                trans_file.seek(0)
                wb = openpyxl.load_workbook(trans_file)
                ws = wb[sheet_trans]

                # Destination Column Index සොයාගැනීම (Header Row එක මත පදනම්ව)
                header_idx = header_row_trans + 1
                col_map = {}
                for cell in ws[header_idx]:
                    if cell.value in mapping_dict.values():
                        source_col_name = [k for k, v in mapping_dict.items() if v == cell.value][0]
                        col_map[cell.column] = source_col_name

                # පැරණි දත්ත මකා දැමීම (Formatting මකා නොදමා)
                for row in ws.iter_rows(min_row=header_idx + 1, max_row=ws.max_row):
                    for cell in row:
                        if cell.column in col_map: # තෝරාගත් columns පමණක් හිස් කරයි
                            cell.value = None

                # නව දත්ත පේළියෙන් පේළිය ඇතුළත් කිරීම
                for r_idx, row_data in enumerate(active_df.to_dict('records'), start=header_idx + 1):
                    for c_idx, s_col_name in col_map.items():
                        ws.cell(row=r_idx, column=c_idx).value = row_data[s_col_name]

                output = BytesIO()
                wb.save(output)
                
                st.success("සාර්ථකයි! Formatting සහ Colours ආරක්ෂිතව පවතී.")
                st.download_button("📥 Download Excel", output.getvalue(), "Updated_Data.xlsx")
            except Exception as e:
                st.error(f"Error: {e}")
        else:
            st.warning("කරුණාකර අවම වශයෙන් එක් Column එකක්වත් තෝරන්න.")
